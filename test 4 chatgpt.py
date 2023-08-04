import logging
import asyncio
import openpyxl
import os
from aiogram import Bot, Dispatcher, types

# Set your bot token here
BOT_TOKEN = '6145169134:AAF39yuQRrF8q8xWX0V3cQrrEOuFMBB28GQ'

# Construct the full file path using os.path.join
#file_dir = r"C:\Users\budag\OneDrive\Рабочий стол\Phone Shop"
file_name = "Product List.xlsx"
#EXCEL_FILE_PATH = os.path.join(file_dir, file_name)
EXCEL_FILE_PATH = os.path.join(file_name)

# Create bot and dispatcher instances
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(bot)
logging.basicConfig(level=logging.INFO)

# Dictionary to keep track of user states
user_states = {}


@dp.message_handler(commands=['start'])
async def start(message: types.Message):
    user_id = message.from_user.id
    user_states[user_id] = {}

    await message.answer("Hello! Welcome to the Shop Bot. What do you want to see?",
                         reply_markup=types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True))

    # Load Excel data using openpyxl
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=True)
    sheet_main = workbook['Main']
    product_types = [cell.value for cell in sheet_main['A'][1:] if cell.value.strip()]  # Filter out empty values

    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    keyboard.add(*product_types)

    await message.answer("Please select a product type:", reply_markup=keyboard)


@dp.message_handler(lambda message: message.text in [cell.value for cell in
                                                     openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=True)['Main'][
                                                         'A'][1:] if cell.value.strip()])
async def select_product_type(message: types.Message):
    user_id = message.from_user.id
    user_states[user_id]['product_type'] = message.text

    # Load Excel data using openpyxl
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=True)
    product_type_sheet = user_states[user_id]['product_type']
    sheet_brand = workbook[product_type_sheet]

    brands = list(set([cell.value for cell in sheet_brand['B'][1:] if cell.value.strip()]))[
             ::-1]  # Get unique brand names and reverse the list
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    keyboard.add(*brands)
    ##

    await message.answer("Please select a brand:", reply_markup=keyboard)


@dp.message_handler(lambda message: message.text in list(set([cell.value for cell in
                                                              openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=True)[
                                                                  user_states[message.from_user.id]['product_type']][
                                                                  'B'][1:] if cell.value.strip()])))
async def select_brand(message: types.Message):
    user_id = message.from_user.id
    user_states[user_id]['brand'] = message.text

    # Load Excel data using openpyxl
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=True)
    product_type_sheet = user_states[user_id]['product_type']
    brand = user_states[user_id]['brand']  # Store selected brand in user state
    sheet_model = workbook[product_type_sheet]

    models = list(set([cell.value for cell in sheet_model['A'] if
                       sheet_model['B'][cell.row - 1].value == brand]))  # Get unique model names

    # Filter models that have "Available" value as "Yes"
    available_models = []
    for row in sheet_model.iter_rows(min_row=2, values_only=True):
        if row[1] == brand and row[8] == "Yes":
            available_models.append(row[0])

    models = list(set(models) & set(available_models))

    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    keyboard.add(*models)

    await message.answer("Please select a model:", reply_markup=keyboard)


@dp.message_handler(lambda message: message.text in list(set([cell.value for cell in
                                                              openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=True)[
                                                                  user_states[message.from_user.id]['product_type']][
                                                                  'A'][1:] if cell.value.strip()])))
async def select_model(message: types.Message):
    user_id = message.from_user.id
    user_states[user_id]['model'] = message.text

    # Load Excel data using openpyxl
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=True)
    product_type_sheet = user_states[user_id]['product_type']
    brand = user_states[user_id]['brand']  # Retrieve selected brand from user state
    model = user_states[user_id]['model']
    sheet_model = workbook[product_type_sheet]

    model_data = None
    for row in sheet_model.iter_rows(min_row=2, values_only=True):
        if row[0] == model and row[1] == brand and row[8] == "Yes":  # Check for model, brand, and "Available" column
            model_data = row  # Store the entire row as model data
            break

    if not model_data:
        await message.answer("This model is not available.")
        return

    description = model_data[4]
    image = model_data[5]
    price = model_data[7]
    link = model_data[6]

    colors = set()
    memories = set()
    for row in sheet_model.iter_rows(min_row=2, values_only=True):
        if row[0] == model and row[1] == brand and row[8] == "Yes":
            colors.add(row[3])
            memories.add(str(row[2]))  # Convert to string

    # Display memory and color on separate rows
    info_text = f"Brand: {brand}\nModel: {model}\nDescription: {description}\nMemory: {'/'.join(memories)}\nColor: {'/'.join(colors)}\nPrice: {price}"

    keyboard = types.InlineKeyboardMarkup()
    keyboard.add(types.InlineKeyboardButton("Buy Now", url=link))
    keyboard.add(types.InlineKeyboardButton("Restart", callback_data="restart"))

    await message.answer_photo(photo=image, caption=info_text, reply_markup=keyboard)


@dp.callback_query_handler(lambda callback_query: callback_query.data == 'restart')
async def restart(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id
    await bot.send_message(user_id, "Bot has been restarted.")
    await start(callback_query.message)


@dp.message_handler()
async def unknown(message: types.Message):
    await message.answer("Sorry, I didn't understand that.")


if __name__ == '__main__':
    from aiogram import executor

    loop = asyncio.get_event_loop()
    executor.start_polling(dp, loop=loop, skip_updates=True)