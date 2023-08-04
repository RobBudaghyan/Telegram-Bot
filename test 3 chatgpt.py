import asyncio
import logging
import openpyxl
from aiogram import Bot, Dispatcher, types
from aiogram.contrib.middlewares.logging import LoggingMiddleware
from aiogram.types import ParseMode
import os

# Set up your Telegram Bot Token here
TOKEN = "6145169134:AAF39yuQRrF8q8xWX0V3cQrrEOuFMBB28GQ"
filePath = os.path.join(os.getcwd(), r"C:\Users\budag\OneDrive\Рабочий стол\Phone Shop\Product List.xlsx")

bot = Bot(token=TOKEN)
dp = Dispatcher(bot)
logging.basicConfig(level=logging.INFO)
dp.middleware.setup(LoggingMiddleware())


@dp.message_handler(commands=['start'])
async def send_welcome(message: types.Message):
    welcome_message = "Welcome to our shop! 👋\nWhat do you want to see?"
    await message.reply(welcome_message, reply_markup=await get_catalog_buttons())


async def get_catalog_buttons():
    workbook = openpyxl.load_workbook(filePath, data_only=True)
    sheet_main = workbook['Main']
    catalog_options = [cell.value for cell in sheet_main['A']]
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    keyboard.add(*(types.KeyboardButton(text) for text in catalog_options))
    return keyboard


@dp.message_handler(lambda message: message.text in ["Buy now!", "Restart"])
async def process_actions(message: types.Message):
    if message.text == "Buy now!":
        await message.answer("Click the link to buy the product.", disable_web_page_preview=False)
    elif message.text == "Restart":
        await send_welcome(message)


async def process_model_selection(message: types.Message, selected_brand):
    selected_model = message.text
    workbook = openpyxl.load_workbook(filePath, data_only=True)
    sheet_brand = workbook['Smartphones']
    model_data = None
    for row in sheet_brand.iter_rows(min_row=2, values_only=True):
        if row[1] == selected_brand and row[0] == selected_model and row[8] == "Yes":
            model_data = {
                "brand": row[1],
                "model_name": row[0],
                "memory": row[2],
                "color": row[3],
                "description": row[4],
                "image_link": row[5],
                "buy_link": row[6],
                "price": row[7]
            }
            break

    if model_data:
        memory_options = [row[2] for row in sheet_brand.iter_rows(min_row=1, values_only=True) if
                          row[1] == selected_brand and row[0] == selected_model and row[8] == "Yes"]
        unique_memory_options = list(set(memory_options))
        memory_str = "/".join(unique_memory_options)
        detailed_message = (
            f"{model_data['brand']} {model_data['model_name']}\n"
            f"Description: {model_data['description']}\n"
            f"Memory: {memory_str}\n"
            f"Color: {model_data['color']}\n"
            f"Price: {model_data['price']}\n"
            f"Buy now: {model_data['buy_link']}"
        )
        await message.answer_photo(model_data['image_link'], detailed_message, parse_mode=ParseMode.MARKDOWN)
        await message.answer("Choose an action:", reply_markup=await get_action_buttons())


async def get_brand_buttons(selected_category):
    workbook = openpyxl.load_workbook(filePath, data_only=True)
    sheet_category = workbook[selected_category]
    brand_options = [cell.value for cell in sheet_category['B'] if cell.row > 1]
    unique_brands = list(set(brand_options))
    return unique_brands


async def check_message_in_catalog_buttons(text):
    return text in [btn.text for btn in await get_catalog_buttons()]


@dp.message_handler(lambda message: check_message_in_catalog_buttons(message.text))
async def process_category_selection(message: types.Message):
    selected_category = message.text
    workbook = openpyxl.load_workbook(filePath, data_only=True)
    sheet_category = workbook[selected_category]
    brand_options = [cell.value for cell in sheet_category['B'] if cell.row > 1]
    unique_brands = list(set(brand_options))
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    keyboard.add(*(types.KeyboardButton(text) for text in unique_brands))
    await message.answer("Select a brand:", reply_markup=keyboard)


@dp.message_handler(lambda message: message.text in asyncio.ensure_future(get_brand_buttons(message.text)))
async def process_model_selection_wrapper(message: types.Message):
    selected_brand = message.text
    workbook = openpyxl.load_workbook(filePath, data_only=True)
    sheet_brand = workbook['Smartphones']
    model_names = [row[0] for row in sheet_brand.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True) if
                   row[1] == selected_brand and row[8] == "Yes"]
    unique_model_names = list(set(model_names))
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    keyboard.add(*(types.KeyboardButton(model_name) for model_name in unique_model_names))
    await message.answer("Select a model:", reply_markup=keyboard)


async def get_action_buttons():
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    keyboard.add(types.KeyboardButton("Buy now!"), types.KeyboardButton("Restart"))
    return keyboard


if __name__ == '__main__':
    from aiogram import executor

    executor.start_polling(dp, skip_updates=True)
